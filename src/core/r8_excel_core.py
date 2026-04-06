# src/core/r8_excel_core.py
from __future__ import annotations
from pathlib import Path
from typing import Dict, List
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

AZUL_OSCURO  = "1F3864"
AZUL_MEDIO   = "2E75B6"
AZUL_CLARO   = "D6E4F0"
BLANCO       = "FFFFFF"
AMARILLO     = "FFF9C4"
AMARILLO_SUAVE = "FFF2CC"

HDR_FILL_PRI   = PatternFill("solid", fgColor=AZUL_OSCURO)
HDR_FILL_PAG   = PatternFill("solid", fgColor="1A5276")
HDR_FILL_RES   = PatternFill("solid", fgColor=AZUL_OSCURO)
ROW_ALT_FILL   = PatternFill("solid", fgColor=AZUL_CLARO)
TOTAL_FILL     = PatternFill("solid", fgColor=AMARILLO)
SUBTOTAL_FILL  = PatternFill("solid", fgColor="E8F4FD")
FALLBACK_FILL  = PatternFill("solid", fgColor=AMARILLO_SUAVE)

HDR_FONT      = Font(bold=True, color=BLANCO, name="Calibri", size=10)
HDR_FONT_DARK = Font(bold=True, color=BLANCO, name="Calibri", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_FONT     = Font(bold=True, name="Calibri", size=10)
TOTAL_FONT    = Font(bold=True, name="Calibri", size=10, color="7B3F00")

CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT    = Alignment(horizontal="left",   vertical="center")
RIGHT   = Alignment(horizontal="right",  vertical="center")

FMT_MONEDA = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
FMT_DATE   = "DD/MM/YYYY"

NS33  = {"cfdi": "http://www.sat.gob.mx/cfd/3", "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital"}
NS40  = {"cfdi": "http://www.sat.gob.mx/cfd/4", "tfd": "http://www.sat.gob.mx/TimbreFiscalDigital"}
NS_P10 = "http://www.sat.gob.mx/Pagos"
NS_P20 = "http://www.sat.gob.mx/Pagos20"

REGIMENES: Dict[str, str] = {
    "601": "General de Ley Personas Morales",
    "603": "Personas Morales con Fines no Lucrativos",
    "605": "Sueldos y Salarios e Ingresos Asimilados a Salarios",
    "606": "Arrendamiento",
    "607": "Régimen de Enajenación o Adquisición de Bienes",
    "608": "Demás Ingresos",
    "609": "Consolidación",
    "610": "Residentes en el Extranjero sin Establecimiento Permanente en México",
    "611": "Ingresos por Dividendos (Socios y Accionistas)",
    "612": "Personas Físicas con Actividades Empresariales y Profesionales",
    "614": "Ingresos por Intereses",
    "615": "Régimen de los ingresos por obtención de premios",
    "616": "Sin obligaciones fiscales",
    "620": "Sociedades Cooperativas de Producción que optan por diferir sus ingresos",
    "621": "Incorporación Fiscal",
    "622": "Actividades Agrícolas, Ganaderas, Silvícolas y Pesqueras",
    "623": "Opcional para Grupos de Sociedades",
    "624": "Coordinados",
    "625": "Régimen de las Actividades Empresariales con ingresos a través de Plataformas Tecnológicas",
    "626": "Régimen Simplificado de Confianza (RESICO)",
    "628": "Hidrocarburos",
    "629": "De los Regímenes Fiscales Preferentes y de las Empresas Multinacionales",
    "630": "Enajenación de acciones en bolsa de valores",
}

FORMAS_PAGO: Dict[str, str] = {
    "01": "Efectivo", "02": "Cheque nominativo", "03": "Transferencia electrónica",
    "04": "Tarjeta de crédito", "05": "Monedero electrónico",
    "06": "Dinero electrónico", "08": "Vales de despensa",
    "12": "Dación en pago", "13": "Pago por subrogación",
    "14": "Pago por consignación", "15": "Condonación",
    "17": "Compensación", "23": "Novación", "24": "Confusión",
    "25": "Remisión de deuda", "26": "Prescripción o caducidad",
    "27": "A satisfacción del acreedor", "28": "Tarjeta de débito",
    "29": "Tarjeta de servicios", "30": "Aplicación de anticipos",
    "31": "Intermediario pagos", "99": "Por definir",
}

TIPOS_COMPROB: Dict[str, str] = {
    "I": "Ingreso", "E": "Egreso", "T": "Traslado",
    "N": "Nómina", "P": "Pago",
}

COLUMNS = [
    "RFC_EMISOR", "RFC_RECEPTOR", "REGIMEN_CODIGO", "REGIMEN_DESC",
    "UUID", "FECHA", "VERSION", "TIPO_COMPROB", "TIPO_COMPROB_DESC",
    "SERIE", "FOLIO", "MONEDA", "TIPO_CAMBIO",
    "SUBTOTAL", "DESCUENTO", "SUBTOTAL_MXN", "TOTAL_C_D",
    "IVA_16_CALC", "TOTAL", "METODO_PAGO", "FORMA_PAGO", "FORMA_PAGO_DESC",
    "EMISOR_NOMBRE", "RECEPTOR_NOMBRE",
]

COLUMNS_PAGOS = [
    "UUID_PAGO", "FOLIO_FAC", "FECHA_PAGO", "FORMA_PAGO_COD", "FORMA_PAGO_DESC",
    "MONEDA_PAGO", "MONTO_PAGADO",
    "UUID_FACTURA_RELACIONADA", "SERIE_FAC",
    "PARCIALIDAD", "SALDO_ANTERIOR", "IMPORTE_PAGADO", "SALDO_INSOLUTO",
    "NOMBRE_CLIENTE", "RFC_CLIENTE",
    "SUBTOTAL_MXN", "IVA_16_CALC", "TOTAL",
    "CLIENTE_FALLBACK",
]

def _sf(x, default=0.0) -> float:
    try:
        if x in (None, ""):
            return float(default)
        return float(str(x).replace(",", ""))
    except Exception:
        return float(default)

def _to_date(fecha: str) -> str:
    return (fecha or "")[:10]

def _ns(root) -> dict:
    return NS33 if "cfd/3" in root.tag else NS40

def _ga(elem, *names, default="") -> str:
    for n in names:
        v = elem.get(n)
        if v is not None:
            return v
    return default

def _norm_uuid(x) -> str:
    return str(x or "").strip().upper()

def parse_cfdi_bytes(xml_bytes: bytes) -> Dict:
    root = ET.fromstring(xml_bytes)
    ns = _ns(root)

    version  = _ga(root, "Version", "version")
    fecha    = _to_date(_ga(root, "Fecha", "fecha"))
    tipo     = (_ga(root, "TipoDeComprobante", "tipoDeComprobante") or "").upper()
    serie    = _ga(root, "Serie", "serie")
    folio    = _ga(root, "Folio", "folio")
    moneda   = (_ga(root, "Moneda", "moneda") or "MXN").upper()
    tc       = _ga(root, "TipoCambio", "TipoDeCambio", "tipoCambio")
    subtotal = _ga(root, "SubTotal", "subTotal", "Subtotal", "subtotal")
    desc     = _ga(root, "Descuento", "descuento", default="0")
    mpago    = _ga(root, "MetodoPago", "metodoPago")
    fpago    = _ga(root, "FormaPago", "formaPago")

    em = root.find("cfdi:Emisor", ns)
    rc = root.find("cfdi:Receptor", ns)

    rfc_em = (em.get("Rfc") or em.get("rfc") or "") if em is not None else ""
    nom_em = (em.get("Nombre") or em.get("nombre") or "") if em is not None else ""
    rfc_rc = (rc.get("Rfc") or rc.get("rfc") or "") if rc is not None else ""
    nom_rc = (rc.get("Nombre") or rc.get("nombre") or "") if rc is not None else ""

    reg_cod = ""
    if rc is not None:
        reg_cod = (rc.get("RegimenFiscalReceptor") or rc.get("regimenFiscalReceptor") or "")
    reg_desc = REGIMENES.get(reg_cod, reg_cod)

    tfd = root.find("cfdi:Complemento/tfd:TimbreFiscalDigital", ns)
    uuid = _norm_uuid((tfd.get("UUID") if tfd is not None else "") or "")

    tc_f     = 1.0 if moneda in ("", "MXN", "XXX") else _sf(tc, 1.0)
    sub_f    = _sf(subtotal)
    dsc_f    = _sf(desc)
    sub_mxn  = sub_f * (tc_f if moneda not in ("MXN", "XXX", "") else 1.0)
    dsc_mxn  = dsc_f * (tc_f if moneda not in ("MXN", "XXX", "") else 1.0)
    total_cd = sub_mxn - dsc_mxn
    iva      = round(total_cd * 0.16, 2)
    total    = round(total_cd + iva, 2)

    return {
        "RFC_EMISOR": rfc_em,
        "RFC_RECEPTOR": rfc_rc,
        "REGIMEN_CODIGO": reg_cod,
        "REGIMEN_DESC": reg_desc,
        "UUID": uuid,
        "FECHA": fecha,
        "VERSION": version,
        "TIPO_COMPROB": tipo,
        "TIPO_COMPROB_DESC": TIPOS_COMPROB.get(tipo, tipo),
        "SERIE": serie,
        "FOLIO": folio,
        "MONEDA": moneda,
        "TIPO_CAMBIO": tc_f if moneda not in ("MXN", "XXX", "") else 1.0,
        "SUBTOTAL": round(sub_f, 2),
        "DESCUENTO": round(dsc_f, 2),
        "SUBTOTAL_MXN": round(sub_mxn, 2),
        "TOTAL_C_D": round(total_cd, 2),
        "IVA_16_CALC": iva,
        "TOTAL": total,
        "METODO_PAGO": mpago,
        "FORMA_PAGO": fpago,
        "FORMA_PAGO_DESC": FORMAS_PAGO.get(str(fpago).zfill(2) if fpago else "", fpago or ""),
        "EMISOR_NOMBRE": nom_em,
        "RECEPTOR_NOMBRE": nom_rc,
        "_XML_ROOT": root,
        "_NS": ns,
    }

def parse_pagos_from_cfdi(row: Dict, role: str = "RECIBIDAS") -> List[Dict]:
    root = row.get("_XML_ROOT")
    ns   = row.get("_NS", {})
    if root is None or row.get("TIPO_COMPROB") != "P":
        return []

    role = (role or "RECIBIDAS").upper()
    uuid_pago = _norm_uuid(row.get("UUID", ""))

    if role == "EMITIDAS":
        nom_cliente_fallback = row.get("RECEPTOR_NOMBRE", "") or ""
        rfc_cliente_fallback = row.get("RFC_RECEPTOR", "") or ""
    else:
        nom_cliente_fallback = row.get("EMISOR_NOMBRE", "") or ""
        rfc_cliente_fallback = row.get("RFC_EMISOR", "") or ""

    pagos_out: List[Dict] = []

    for ns_pago in (NS_P20, NS_P10):
        complemento = root.find("cfdi:Complemento", ns)
        if complemento is None:
            continue

        pagos_node = complemento.find(f"{{{ns_pago}}}Pagos")
        if pagos_node is None:
            continue

        for pago in pagos_node.findall(f"{{{ns_pago}}}Pago"):
            fecha_pago = _to_date(pago.get("FechaPago", ""))
            moneda_p   = pago.get("MonedaP", "MXN")
            monto      = _sf(pago.get("Monto", "0"))
            fp_cod     = pago.get("FormaDePagoP", pago.get("FormaPagoP", ""))
            fp_desc    = FORMAS_PAGO.get(fp_cod.zfill(2) if fp_cod else "", fp_cod)

            for docto in pago.findall(f"{{{ns_pago}}}DoctoRelacionado"):
                id_doc  = _norm_uuid(docto.get("IdDocumento", ""))
                serie   = docto.get("Serie", "")
                folio   = docto.get("Folio", "")
                num_p   = docto.get("NumParcialidad", "")
                saldo_a = _sf(docto.get("ImpSaldoAnt", "0"))
                imp_pag = _sf(docto.get("ImpPagado", monto))
                saldo_i = _sf(docto.get("ImpSaldoInsoluto", "0"))

                total_pago = round(imp_pag, 2)
                subtotal_pago = round(total_pago / 1.16, 2) if total_pago else 0.0
                iva_pago = round(total_pago - subtotal_pago, 2)

                pagos_out.append({
                    "UUID_PAGO": uuid_pago,
                    "FOLIO_FAC": folio,
                    "FECHA_PAGO": fecha_pago,
                    "FORMA_PAGO_COD": fp_cod,
                    "FORMA_PAGO_DESC": fp_desc,
                    "MONEDA_PAGO": moneda_p,
                    "MONTO_PAGADO": round(monto, 2),
                    "UUID_FACTURA_RELACIONADA": id_doc,
                    "SERIE_FAC": serie,
                    "PARCIALIDAD": num_p,
                    "SALDO_ANTERIOR": round(saldo_a, 2),
                    "IMPORTE_PAGADO": total_pago,
                    "SALDO_INSOLUTO": round(saldo_i, 2),
                    "NOMBRE_CLIENTE": nom_cliente_fallback,
                    "RFC_CLIENTE": rfc_cliente_fallback,
                    "SUBTOTAL_MXN": subtotal_pago,
                    "IVA_16_CALC": iva_pago,
                    "TOTAL": total_pago,
                    "CLIENTE_FALLBACK": 1,
                })
        break

    return pagos_out

def _enrich_pagos_with_related_cfdi(df_pagos: pd.DataFrame, df_all: pd.DataFrame, role: str = "RECIBIDAS") -> pd.DataFrame:
    if df_pagos.empty:
        return df_pagos

    role = (role or "RECIBIDAS").upper()
    out = df_pagos.copy()
    out["UUID_FACTURA_RELACIONADA"] = out["UUID_FACTURA_RELACIONADA"].map(_norm_uuid)

    if df_all.empty:
        return out

    if role == "EMITIDAS":
        rel_map = df_all.loc[:, ["UUID", "FOLIO", "RECEPTOR_NOMBRE", "RFC_RECEPTOR"]].copy()
        rel_map["UUID"] = rel_map["UUID"].map(_norm_uuid)
        rel_map = rel_map.drop_duplicates(subset=["UUID"]).rename(columns={
            "UUID": "UUID_FACTURA_RELACIONADA",
            "FOLIO": "FOLIO_FAC_REL",
            "RECEPTOR_NOMBRE": "NOMBRE_CLIENTE_REL",
            "RFC_RECEPTOR": "RFC_CLIENTE_REL",
        })
    else:
        rel_map = df_all.loc[:, ["UUID", "FOLIO", "EMISOR_NOMBRE", "RFC_EMISOR"]].copy()
        rel_map["UUID"] = rel_map["UUID"].map(_norm_uuid)
        rel_map = rel_map.drop_duplicates(subset=["UUID"]).rename(columns={
            "UUID": "UUID_FACTURA_RELACIONADA",
            "FOLIO": "FOLIO_FAC_REL",
            "EMISOR_NOMBRE": "NOMBRE_CLIENTE_REL",
            "RFC_EMISOR": "RFC_CLIENTE_REL",
        })

    out = out.merge(rel_map, on="UUID_FACTURA_RELACIONADA", how="left")

    match_mask = (
        out["NOMBRE_CLIENTE_REL"].notna()
        & (out["NOMBRE_CLIENTE_REL"].astype(str).str.strip() != "")
    )

    out.loc[match_mask, "FOLIO_FAC"] = out.loc[match_mask, "FOLIO_FAC_REL"]
    out.loc[match_mask, "NOMBRE_CLIENTE"] = out.loc[match_mask, "NOMBRE_CLIENTE_REL"]
    out.loc[match_mask, "RFC_CLIENTE"] = out.loc[match_mask, "RFC_CLIENTE_REL"]
    out.loc[match_mask, "CLIENTE_FALLBACK"] = 0

    out = out.drop(columns=["FOLIO_FAC_REL", "NOMBRE_CLIENTE_REL", "RFC_CLIENTE_REL"], errors="ignore")
    return out

def _blank_duplicate_monto_pagado(df_pagos: pd.DataFrame) -> pd.DataFrame:
    if df_pagos.empty:
        return df_pagos

    out = df_pagos.sort_values(
        ["FECHA_PAGO", "UUID_PAGO", "UUID_FACTURA_RELACIONADA"],
        na_position="last",
    ).reset_index(drop=True)

    dup_mask = out["UUID_PAGO"].astype(str).duplicated(keep="first")
    out.loc[dup_mask, "MONTO_PAGADO"] = ""
    return out

def build_monthly_excels_from_xml_bytes(xml_blobs: List[bytes], role: str = "RECIBIDAS") -> Dict[str, pd.DataFrame]:
    role = (role or "RECIBIDAS").upper()

    seen: set = set()
    rows: List[Dict] = []
    pagos_rows: List[Dict] = []

    for b in xml_blobs:
        try:
            row = parse_cfdi_bytes(b)
            uuid = _norm_uuid(row.get("UUID", ""))
            if not uuid or uuid in seen:
                continue
            seen.add(uuid)

            pagos_rows.extend(parse_pagos_from_cfdi(row, role=role))

            row.pop("_XML_ROOT", None)
            row.pop("_NS", None)
            rows.append(row)
        except Exception:
            continue

    df_all = pd.DataFrame(rows, columns=COLUMNS) if rows else pd.DataFrame(columns=COLUMNS)
    if not df_all.empty:
        df_all["UUID"] = df_all["UUID"].map(_norm_uuid)
        df_all = (
            df_all.assign(_ord=df_all["FECHA"].astype(str))
            .sort_values(["_ord", "UUID"])
            .drop(columns=["_ord"])
            .reset_index(drop=True)
        )

    df_pue = (
        df_all[df_all["METODO_PAGO"].astype(str).str.upper().eq("PUE")].copy()
        if not df_all.empty else pd.DataFrame(columns=COLUMNS)
    )

    df_pagos = pd.DataFrame(pagos_rows, columns=COLUMNS_PAGOS) if pagos_rows else pd.DataFrame(columns=COLUMNS_PAGOS)
    if not df_pagos.empty:
        df_pagos = _enrich_pagos_with_related_cfdi(df_pagos, df_all, role=role)
        df_pagos = _blank_duplicate_monto_pagado(df_pagos)
        df_pagos = df_pagos.reindex(columns=COLUMNS_PAGOS)

    df_resumen = _build_resumen(df_all, df_pue)

    return {
        "CFDI": df_all,
        "CFDI_PUE": df_pue,
        "PAGOS": df_pagos,
        "Resumen": df_resumen,
    }

def _build_resumen(df: pd.DataFrame, df_pue: pd.DataFrame | None = None) -> pd.DataFrame:
    if df.empty and (df_pue is None or df_pue.empty):
        return pd.DataFrame(columns=["CONCEPTO", "No. CFDI", "SUBTOTAL", "IVA (16%)", "TOTAL"])

    tipos_orden = [
        ("I", "Ingresos (Facturas)"),
        ("N", "Nómina"),
        ("T", "Traslados"),
        ("E", "Egresos / Notas de crédito"),
        ("P", "Complementos de Pago"),
    ]

    filas: List[Dict] = []

    if not df.empty:
        for codigo, label in tipos_orden:
            sub = df[df["TIPO_COMPROB"] == codigo]
            if sub.empty:
                continue
            filas.append({
                "CONCEPTO": label,
                "No. CFDI": len(sub),
                "SUBTOTAL": round(sub["SUBTOTAL_MXN"].sum(), 2),
                "IVA (16%)": round(sub["IVA_16_CALC"].sum(), 2),
                "TOTAL": round(sub["TOTAL"].sum(), 2),
            })

    if df_pue is not None and not df_pue.empty:
        filas.append({
            "CONCEPTO": "PUE",
            "No. CFDI": len(df_pue),
            "SUBTOTAL": round(df_pue["SUBTOTAL_MXN"].sum(), 2),
            "IVA (16%)": round(df_pue["IVA_16_CALC"].sum(), 2),
            "TOTAL": round(df_pue["TOTAL"].sum(), 2),
        })

    total_base = df if not df.empty else pd.DataFrame(columns=["SUBTOTAL_MXN", "IVA_16_CALC", "TOTAL"])
    total_row = {
        "CONCEPTO": "TOTAL GENERAL",
        "No. CFDI": int(len(total_base)),
        "SUBTOTAL": round(total_base["SUBTOTAL_MXN"].sum(), 2) if not total_base.empty else 0.0,
        "IVA (16%)": round(total_base["IVA_16_CALC"].sum(), 2) if not total_base.empty else 0.0,
        "TOTAL": round(total_base["TOTAL"].sum(), 2) if not total_base.empty else 0.0,
    }

    return pd.concat([pd.DataFrame(filas), pd.DataFrame([total_row])], ignore_index=True)

def save_excels_with_format(sheets: Dict[str, pd.DataFrame], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if out_path.exists():
        out_path.unlink()

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        for name in ["CFDI", "CFDI_PUE", "PAGOS", "Resumen"]:
            dfx = sheets.get(name, pd.DataFrame())
            (dfx if not dfx.empty else dfx.head(0)).to_excel(w, sheet_name=name, index=False)

        wb = w.book  # type: ignore
        _format_sheet_cfdi(wb, "CFDI")
        _format_sheet_cfdi(wb, "CFDI_PUE")
        _format_sheet_pagos(wb, "PAGOS")
        _format_sheet_resumen(wb, "Resumen")

    return out_path

def _format_sheet_cfdi(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    cols = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    col_moneda = {"SUBTOTAL", "DESCUENTO", "SUBTOTAL_MXN", "TOTAL_C_D", "IVA_16_CALC", "TOTAL", "TIPO_CAMBIO"}
    col_fecha = {"FECHA"}

    for cell in ws[1]:
        cell.font = HDR_FONT_DARK
        cell.fill = HDR_FILL_PRI
        cell.alignment = CENTER

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ROW_ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, cell in enumerate(row):
            col_name = cols[c_idx] if c_idx < len(cols) else ""
            cell.font = BODY_FONT
            cell.alignment = RIGHT if col_name in col_moneda else CENTER if col_name in col_fecha else LEFT
            if col_name in col_moneda:
                cell.number_format = FMT_MONEDA
            elif col_name in col_fecha:
                cell.number_format = FMT_DATE
            if col_name == "TOTAL":
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
            elif fill:
                cell.fill = fill

    _auto_width(ws, cols)

def _format_sheet_pagos(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    cols = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    col_moneda = {"MONTO_PAGADO", "SALDO_ANTERIOR", "IMPORTE_PAGADO", "SALDO_INSOLUTO", "SUBTOTAL_MXN", "IVA_16_CALC", "TOTAL"}
    col_fecha = {"FECHA_PAGO"}
    idx_flag = cols.index("CLIENTE_FALLBACK") + 1 if "CLIENTE_FALLBACK" in cols else None

    for cell in ws[1]:
        cell.font = HDR_FONT_DARK
        cell.fill = HDR_FILL_PAG
        cell.alignment = CENTER

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ROW_ALT_FILL if r_idx % 2 == 0 else None
        fallback_value = ws.cell(r_idx, idx_flag).value if idx_flag else None

        for c_idx, cell in enumerate(row):
            col_name = cols[c_idx] if c_idx < len(cols) else ""
            cell.font = BODY_FONT
            cell.alignment = RIGHT if col_name in col_moneda else CENTER if col_name in col_fecha else LEFT

            if col_name in col_moneda and cell.value not in (None, ""):
                cell.number_format = FMT_MONEDA
            elif col_name in col_fecha:
                cell.number_format = FMT_DATE

            if col_name == "SALDO_INSOLUTO" and cell.value not in (None, "") and float(cell.value or 0) > 0:
                cell.fill = PatternFill("solid", fgColor="FDEBD0")
            elif col_name == "SALDO_INSOLUTO" and cell.value == 0:
                cell.fill = PatternFill("solid", fgColor="D5F5E3")
            elif col_name == "TOTAL":
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
            elif col_name == "NOMBRE_CLIENTE" and str(fallback_value) == "1":
                cell.fill = FALLBACK_FILL
            elif fill:
                cell.fill = fill

    if idx_flag:
        ws.column_dimensions[get_column_letter(idx_flag)].hidden = True

    _auto_width(ws, cols)

def _format_sheet_resumen(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    cols = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    col_moneda = {"SUBTOTAL", "IVA (16%)", "TOTAL"}

    for cell in ws[1]:
        cell.font = HDR_FONT_DARK
        cell.fill = HDR_FILL_RES
        cell.alignment = CENTER

    ws.freeze_panes = "A2"
    total_row = ws.max_row

    for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        is_total = r_idx == total_row
        for c_idx, cell in enumerate(row):
            col_name = cols[c_idx] if c_idx < len(cols) else ""
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
            else:
                cell.font = BOLD_FONT if col_name == "CONCEPTO" else BODY_FONT
                cell.fill = SUBTOTAL_FILL if r_idx % 2 == 0 else PatternFill()
            cell.alignment = RIGHT if col_name in col_moneda else LEFT
            if col_name in col_moneda:
                cell.number_format = FMT_MONEDA

    _auto_width(ws, cols)

def _auto_width(ws, cols: list):
    col_anchos_min = {
        "UUID": 38,
        "UUID_FACTURA_RELACIONADA": 38,
        "UUID_PAGO": 38,
        "EMISOR_NOMBRE": 30,
        "RECEPTOR_NOMBRE": 30,
        "NOMBRE_CLIENTE": 30,
        "REGIMEN_DESC": 40,
        "FORMA_PAGO_DESC": 25,
        "TIPO_COMPROB_DESC": 20,
        "CONCEPTO": 35,
    }
    for c_idx, col_name in enumerate(cols, 1):
        col_letter = get_column_letter(c_idx)
        min_w = col_anchos_min.get(col_name, 10)
        col_data = ws[col_letter]
        max_len = max((len(str(cell.value)) if cell.value is not None else 0 for cell in col_data), default=min_w)
        ws.column_dimensions[col_letter].width = min(max(min_w, max_len + 2), 50)